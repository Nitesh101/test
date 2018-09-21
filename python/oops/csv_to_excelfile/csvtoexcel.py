import csv
from openpyxl import Workbook
from openpyxl.cell import get_column_letter

f = open(r'/home/madisnit/Desktop/oops/csv_to_excelfile/csv_to_excel.csv')

csv.register_dialect('colons', delimiter=':')

reader = csv.reader(f, dialect='colons')

wb = Workbook()
dest_filename = r"/home/madisnit/Desktop/oops/csv_to_excelfile/herp.xlsx"

ws = wb.worksheets[0]
ws.title = "A Snazzy Title"

for row_index, row in enumerate(reader):
    for column_index, cell in enumerate(row):
        column_letter = get_column_letter((column_index + 1))
        ws.cell('%s%s'%(column_letter, (row_index + 1))).value = cell

wb.save(filename = dest_filename)
